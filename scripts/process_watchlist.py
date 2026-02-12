#!/usr/bin/env python3
"""Read concatenated messages from stdin and write them into an Excel file.
Usage: osascript extract_mail.scpt | python3 process_watchlist.py --out /path/to/watchlist.xlsx
"""
import sys
import argparse
import pandas as pd
import io
import re
import csv
from datetime import datetime
import os

SEPARATOR = '---MSG---'

def parse_message(msg_text):
    """Parse message to extract recap date and tabular data."""
    lines = [l.rstrip('\r') for l in msg_text.splitlines()]
    if not lines:
        return None, pd.DataFrame()
    
    # Extract the date from "End-of-Day Recap - Price quotes for Tue, January 27, 2026"
    recap_date = None
    data_start_idx = 0
    for i, line in enumerate(lines):
        if 'End-of-Day Recap' in line and 'Price quotes for' in line:
            # Extract date part: "Tue, January 27, 2026"
            match = re.search(r'Price quotes for (.+?)(?:\n|$)', line)
            if match:
                recap_date = match.group(1).strip()
            data_start_idx = i + 1
            break
    
    # Find the header row (contains "Symbol", "Latest", etc.)
    header_idx = None
    for i in range(data_start_idx, len(lines)):
        if lines[i].strip() == 'Symbol':
            header_idx = i
            break
    
    if header_idx is None:
        # No header found; try to parse as-is
        return recap_date, pd.DataFrame()
    
    # Collect header from consecutive lines after "Symbol"
    header = ['Symbol']
    data_start_line = header_idx + 1
    
    # Collect header columns (they appear as separate lines)
    i = header_idx + 1
    while i < len(lines) and lines[i].strip() and not lines[i][0].isdigit() and not lines[i][0] in ('^', '$', 'N'):
        col = lines[i].strip()
        if col and col not in header:
            header.append(col)
        else:
            break
        i += 1
    
    # Data rows start after header columns
    data_start_line = i
    
    # Collect data rows: each row is N consecutive lines where N = len(header)
    rows = []
    current_row = []
    for i in range(data_start_line, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        
        # If we have enough columns for a row, start a new one
        if len(current_row) == len(header):
            rows.append(current_row)
            current_row = []
        
        current_row.append(line)
        
        # Stop after ^USDCHF row is complete
        if len(current_row) == len(header) and current_row[0] == '^USDCHF':
            rows.append(current_row)
            break
    
    # Add the last row if it's complete and not already added
    if current_row and len(current_row) == len(header) and (not rows or current_row != rows[-1]):
        rows.append(current_row)
    
    # Create DataFrame with header
    if rows:
        try:
            print("Parsed rows:", file=sys.stderr)
            for r in rows:
                print(r, file=sys.stderr)
            df = pd.DataFrame(rows, columns=header)
            # Clean symbol names: remove $ and ^ characters
            if 'Symbol' in df.columns:
                df['Symbol'] = df['Symbol'].str.replace(r'[\$\^]', '', regex=True)
            # Convert numeric columns to float
            numeric_cols = ['Latest', 'Change', 'Open', 'High', 'Low', 'Volume']
            for col in numeric_cols:
                if col in df.columns:
                    print(f"Raw {col} values: ", df[col].tolist(), file=sys.stderr)
                    raw = df[col].astype(str)
                    unch_mask = raw.str.strip().str.match(r'^unch\w*$', case=False, na=False)
                    df[col] = pd.to_numeric(
                        raw
                        .str.replace(',', '', regex=False)
                        .str.replace('s', '', regex=False)
                        .str.replace('+', '', regex=False)
                        .str.replace(r'^\s*unch\w*\s*$', '0', regex=True, case=False),
                        errors='coerce'
                    )
                    df.loc[unch_mask, col] = 0.0
                    print(f"Converted {col} values: ", df[col].tolist(), file=sys.stderr)
            # Convert %Change: remove % prefix and convert to float (will be formatted as percentage in Excel)
            if '%Change' in df.columns:
                print("Raw %Change values: ", df['%Change'].tolist(), file=sys.stderr)
                raw_pct = df['%Change'].astype(str)
                unch_mask = raw_pct.str.strip().str.match(r'^unch\w*$', case=False, na=False)
                df['%Change'] = pd.to_numeric(
                    raw_pct
                    .str.replace('%', '', regex=False)
                    .str.replace(',', '', regex=False)
                    .str.replace('+', '', regex=False)
                    .str.replace(r'^\s*unch\w*\s*$', '0', regex=True, case=False),
                    errors='coerce'
                )
                # Keep display values consistent with percent-based inputs.
                abs_pct = df['%Change'].abs()
                df.loc[abs_pct > 1, '%Change'] = df.loc[abs_pct > 1, '%Change'] / 100.0
                df.loc[unch_mask, '%Change'] = 0.0
                print("Converted %Change values: ", df['%Change'].tolist(), file=sys.stderr)
        except Exception as e:
            print(f"Conversion error: {e}", file=sys.stderr)
            df = pd.DataFrame(rows)
    else:
        df = pd.DataFrame()
    
    return recap_date, df

def _get_existing_dates(ws):
    """Extract the set of date strings already present in a worksheet.
    
    Each date block has a 'Time' column (last in the block) whose data cells
    contain the date string (e.g. '02/03/26').  Walk the header row to find
    every 'Time' column, then read the first data cell beneath it.
    """
    existing = set()
    for col in range(1, (ws.max_column or 0) + 1):
        if ws.cell(row=1, column=col).value == 'Time':
            val = ws.cell(row=2, column=col).value
            if val:
                existing.add(str(val).strip())
    return existing


def _next_free_col(ws):
    """Return the first completely empty column (1-based) in *ws*."""
    max_col = ws.max_column or 0
    # walk backwards to skip trailing None-only columns openpyxl may report
    for c in range(max_col, 0, -1):
        for r in range(1, (ws.max_row or 0) + 1):
            if ws.cell(row=r, column=c).value is not None:
                return c + 3  # 2-column gap after last data column
    return 1


def _write_date_block(ws, current_col, df):
    """Write a single date block (header + data rows) starting at *current_col*.
    Returns the next available column after the block (with 2-col gap)."""
    # Write header
    for col_num, col_name in enumerate(df.columns, start=0):
        cell = ws.cell(row=1, column=current_col + col_num)
        cell.value = col_name

    # Write data rows
    for row_num, row in enumerate(df.values, start=2):
        for col_num, val in enumerate(row):
            cell = ws.cell(row=row_num, column=current_col + col_num)
            cell.value = val

            # Apply number formatting
            col_name = df.columns[col_num]
            if col_name == '%Change':
                cell.number_format = '0.00%'
            elif col_name == 'Time':
                cell.number_format = 'h:mm AM/PM'
            elif col_name in ['Latest', 'Change', 'Open', 'High', 'Low']:
                cell.number_format = '0.00'
            elif col_name == 'Volume':
                cell.number_format = '0'

    return current_col + len(df.columns) + 2


def write_to_excel_by_month(dfs_with_dates, out_path):
    """Create / update separate Excel files per month, with worksheets per calendar week.
    
    Merges with existing xlsx files so data from previous runs is preserved and
    new dates are appended without duplicates.
    """
    from openpyxl import Workbook, load_workbook
    
    # Get base directory from out_path
    base_dir = os.path.dirname(out_path)
    os.makedirs(base_dir, exist_ok=True)
    
    # Group dataframes by month and calendar week
    month_data = {}  # {month_name: {week_num: [(df, date_str, datetime), ...]}}
    
    for df, recap_date in dfs_with_dates:
        if not recap_date or df.empty:
            continue
        
        try:
            dt = datetime.strptime(recap_date, '%a, %B %d, %Y')
            month_name = dt.strftime('%B')  # e.g., 'January'
            day = dt.day
            # Calculate calendar week: days 1-7 = Week1, 8-14 = Week2, 15-21 = Week3, 22-28 = Week4, 29+ = Week5
            week_num = (day - 1) // 7 + 1
            
            if month_name not in month_data:
                month_data[month_name] = {}
            if week_num not in month_data[month_name]:
                month_data[month_name][week_num] = []
            month_data[month_name][week_num].append((df, recap_date, dt))
        except Exception as e:
            print(f"Date parse error for '{recap_date}': {e}", file=sys.stderr)
            continue
    
    # Create / update separate Excel file for each month
    for month_name in sorted(month_data.keys()):
        month_file = os.path.join(base_dir, f'{month_name.lower()}.xlsx')
        
        # Load existing workbook or create a new one
        if os.path.exists(month_file):
            wb = load_workbook(month_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
        
        week_nums = sorted(month_data[month_name].keys())
        
        for week_num in week_nums:
            sheet_name = f'Week{week_num}'
            
            # Get or create the sheet
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                existing_dates = _get_existing_dates(ws)
                current_col = _next_free_col(ws)
            else:
                ws = wb.create_sheet(sheet_name)
                existing_dates = set()
                current_col = 1
            
            dfs_for_week = month_data[month_name][week_num]
            dfs_for_week_sorted = sorted(dfs_for_week, key=lambda x: x[2])
            
            added = 0
            for df, recap_date, dt in dfs_for_week_sorted:
                if df.empty:
                    continue
                
                # Derive the short date key that appears in the Time column
                # (e.g. '02/09/26') so we can skip dates already written
                time_col_vals = df['Time'].astype(str).tolist() if 'Time' in df.columns else []
                date_key = time_col_vals[0] if time_col_vals else dt.strftime('%m/%d/%y')
                
                if date_key in existing_dates:
                    continue  # already in the sheet from a previous run
                
                current_col = _write_date_block(ws, current_col, df)
                existing_dates.add(date_key)
                added += 1
            
            if added:
                print(f'  {sheet_name}: added {added} new date(s)', file=sys.stderr)
        
        # Sort sheets by week number
        desired_order = sorted(wb.sheetnames, key=lambda s: int(s.replace('Week', '')) if s.startswith('Week') else 999)
        for i, name in enumerate(desired_order):
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))
        
        wb.save(month_file)
        print(f'Saved {month_file}', file=sys.stderr)

def write_to_excel(dfs_with_dates, out_path):
    # each dataframe in dfs is written to its own sheet named with the recap date
    from openpyxl.styles import numbers
    
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    mode = 'a' if os.path.exists(out_path) else 'w'
    try:
        with pd.ExcelWriter(out_path, engine='openpyxl', mode=mode) as writer:
            for df, sheet_name_base in dfs_with_dates:
                # Use recap_date as sheet name, fallback to timestamp if not available
                if sheet_name_base:
                    sheet_name = sheet_name_base[:31]  # Excel sheet name max 31 chars
                else:
                    sheet_name = 'Watchlist_' + datetime.now().strftime('%Y%m%d_%H%M%S')
                    sheet_name = sheet_name[:31]
                
                if df.empty:
                    pd.DataFrame({'message': ['(empty)']}).to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply formatting to columns
                    ws = writer.sheets[sheet_name]
                    for col_num, col_name in enumerate(df.columns, start=1):
                        col_letter = ws.cell(row=1, column=col_num).column_letter
                        
                        if col_name == '%Change':
                            # Percentage format
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = '0.00%'
                        elif col_name == 'Time':
                            # Time format
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = 'h:mm AM/PM'
                        elif col_name in ['Latest', 'Change', 'Open', 'High', 'Low', 'Volume']:
                            # Number format (2 decimal places for most, 0 for Volume)
                            fmt = '0.00' if col_name != 'Volume' else '0'
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = fmt
    except Exception as e:
        # fallback: try writing single sheet if append mode fails
        with pd.ExcelWriter(out_path, engine='openpyxl', mode='w') as writer:
            for df, sheet_name_base in dfs_with_dates:
                if sheet_name_base:
                    sheet_name = sheet_name_base[:31]
                else:
                    sheet_name = 'Watchlist_' + datetime.now().strftime('%Y%m%d_%H%M%S')
                    sheet_name = sheet_name[:31]
                
                if df.empty:
                    pd.DataFrame({'message': ['(empty)']}).to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply formatting to columns
                    ws = writer.sheets[sheet_name]
                    for col_num, col_name in enumerate(df.columns, start=1):
                        col_letter = ws.cell(row=1, column=col_num).column_letter
                        
                        if col_name == '%Change':
                            # Percentage format
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = '0.00%'
                        elif col_name == 'Time':
                            # Time format
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = 'h:mm AM/PM'
                        elif col_name in ['Latest', 'Change', 'Open', 'High', 'Low', 'Volume']:
                            # Number format (2 decimal places for most, 0 for Volume)
                            fmt = '0.00' if col_name != 'Volume' else '0'
                            for row_num in range(2, len(df) + 2):
                                ws[f'{col_letter}{row_num}'].number_format = fmt

def main():
    parser = argparse.ArgumentParser(description='Process watchlist messages from stdin and write to Excel')
    parser.add_argument('--out', '-o', default=os.path.expanduser('~/Documents/watchlist_summary.xlsx'), help='Output Excel path')
    args = parser.parse_args()

    data = sys.stdin.read()
    if not data.strip():
        print('No input received; no messages found.', file=sys.stderr)
        return

    parts = [p.strip() for p in data.split(SEPARATOR) if p.strip()]
    if not parts:
        print('No messages after splitting; exiting.', file=sys.stderr)
        return

    dfs = []
    for p in parts:
        recap_date, df = parse_message(p)
        dfs.append((df, recap_date))

    write_to_excel_by_month(dfs, args.out)
    print(f'Wrote {len(dfs)} message(s) to {args.out}', file=sys.stderr)

if __name__ == '__main__':
    main()
